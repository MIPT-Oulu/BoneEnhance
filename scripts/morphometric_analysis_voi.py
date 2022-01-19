"""
Calculate morphometric analysis, applying a specific volume-of-interest.
"""

import os
from pathlib import Path
from openpyxl import load_workbook
from time import time, strftime
from scipy.stats import pearsonr
import pandas as pd
from tqdm import tqdm

import numpy as np
import argparse

from bone_enhance.utilities import load, save, print_orthogonal, threshold, calculate_bvtv
from bone_enhance.inference.thickness_analysis import _local_thickness


if __name__ == '__main__':
    start = time()

    # Prediction path
    path = Path('../../Data/Test_set_(full)/input_interpolated')
    t = strftime(f'%Y_%m_%d_%H_%M')
    savepath = f'../../Data/Test_set_(full)/masks_wacv_new/Results_conventional{t}.xlsx'
    snaps = os.listdir(path)
    snaps = [snap for snap in snaps if os.path.isdir(os.path.join(path, snap))]

    correlations = {'Snapshot': [], 'BVTV': [], 'Tb.Th': [], 'Tb.Sp': [], 'Tb.N': [],
                    'BVTV (p)': [], 'Tb.Th (p)': [], 'Tb.Sp (p)': [], 'Tb.N (p)': []}
    for snap in snaps:
        # Remove timestamp from snapshot
        if snap[:2] == '2D' or snap[:2] == '3D':
            snap_short = snap
        else:
            snap_short = snap.split('_', 6)[-1]

        # Arguments
        parser = argparse.ArgumentParser()
        parser.add_argument('--masks', type=Path, default=path.parent / 'trabecular_VOI')
        parser.add_argument('--save', type=Path, default=path.parent / 'masks_wacv_new' / snap_short)
        parser.add_argument('--preds', type=Path, default=path / snap)
        parser.add_argument('--ground_truth', type=Path, default='../../Data/uCT_parameters.csv')
        parser.add_argument('--final_results', type=Path, default='../../Data/final_results.csv')
        parser.add_argument('--plot', type=bool, default=False)
        parser.add_argument('--save_masks', type=bool, default=True)
        parser.add_argument('--batch_id', type=int, default=None)
        parser.add_argument('--resolution', type=tuple, default=(50, 50, 50))  # in µm
        parser.add_argument('--mode', type=str, choices=['med2d_dist3d_lth3d', 'stacked_2d', 'med2d_dist2d_lth3d'],
                            default='med2d_dist3d_lth3d')
        parser.add_argument('--trabecular_number', type=str, choices=['3d', 'plate', 'rod'], default='rod')
        parser.add_argument('--max_th', type=float, default=None)  # in µm

        args = parser.parse_args()

        # Sample list
        samples = os.listdir(args.masks)
        samples.sort()
        if 'visualization' in samples:
            samples.remove('visualization')
        if args.batch_id is not None:
            samples = [samples[args.batch_id]]

        samples_pred = os.listdir(str(args.preds))
        samples_pred.sort()
        if 'visualizations' in samples_pred:
            samples_pred.remove('visualizations')

        # Save paths
        args.save.parent.mkdir(exist_ok=True)
        if args.plot or args.save_masks:
            args.save.mkdir(exist_ok=True)
            (args.save / 'visualization').mkdir(exist_ok=True)

        # Table for results
        results = {'Sample': [],
                   'Trabecular thickness': [], 'Trabecular separation': [], 'BVTV': [], 'Trabecular number': []}

        # Load ground truth values
        target = pd.read_csv(args.ground_truth)

        # Loop for samples
        for idx in tqdm(range(len(samples)), desc=f'Processing snapshot {snap_short}'):
            time_sample = time()
            sample = samples[idx]
            sample_pred = samples_pred[idx]

            # Load prediction and volume-of-interest
            pred, _ = load(str(args.preds / sample_pred / 'conventional_segmentation_gray'), axis=(1, 2, 0,))
            voi, files = load(str(args.masks / sample), axis=(1, 2, 0,))

            if len(np.unique(pred)) != 2:
                pred, _ = threshold(pred)

            if args.plot:
                print_orthogonal(pred, savepath=str(args.save / 'visualization' / (sample + '_pred.png')), res=50 / 1000)

            # Apply volume-of-interest
            pred = np.logical_and(pred, voi).astype(np.uint8) * 255

            # Save binary mask with VOI applied
            if args.save_masks:
                save(str(args.save / sample), Path(sample).stem, pred, dtype='.bmp', verbose=False)

            if args.plot:
                print_orthogonal(pred, savepath=str(args.save / 'visualization' / (sample + '_voi.png')), res=50 / 1000)

            #                       #
            # Morphometric analysis #
            #                       #

            # Thickness map
            th_map = _local_thickness(pred, mode=args.mode, spacing_mm=args.resolution, stack_axis=1,
                                      thickness_max_mm=args.max_th, verbose=False)
            if args.plot:
                print_orthogonal(th_map, cmap='hot', res=args.resolution[0] / 1000)

            # Bone volume fraction
            bvtv = calculate_bvtv(pred, voi) / 100

            # Update results
            th_map = th_map[np.nonzero(th_map)].flatten()
            tb_th = np.mean(th_map)
            results['Sample'].append(sample)
            results['Trabecular thickness'].append(tb_th)
            results['BVTV'].append(bvtv * 100)

            # Separation map
            pred = np.logical_and(np.invert(pred), voi).astype(np.uint8) * 255
            th_map = _local_thickness(pred, mode=args.mode, spacing_mm=args.resolution, stack_axis=1,
                                      thickness_max_mm=args.max_th, verbose=False)
            th_map = th_map[np.nonzero(th_map)].flatten()
            tb_sep = np.mean(th_map)
            results['Trabecular separation'].append(tb_sep)

            # Trabecular number
            if args.trabecular_number == '3d':  # 3D model
                results['Trabecular number'].append(1 / (tb_sep + tb_th))
            elif args.trabecular_number == 'plate':  # 2D plate model
                results['Trabecular number'].append(bvtv / tb_th)
            elif args.trabecular_number == 'rod':  # 2D cylinder rod model
                results['Trabecular number'].append(np.sqrt((4 / np.pi) * bvtv) / tb_th)
            else:  # Append 0 for compatibility
                results['Trabecular number'].append(0)

        #               #
        # Statistics    #
        #               #

        # Pearson correlation

        # Remove NaNs
        results['Trabecular separation'] = list(np.nan_to_num(results['Trabecular separation']))
        results['Trabecular thickness'] = list(np.nan_to_num(results['Trabecular thickness']))
        results['Trabecular number'] = list(np.nan_to_num(results['Trabecular number']))

        # Calculate correlations
        correlations['Snapshot'].append(snap_short)
        correlations['BVTV'].append(pearsonr(results['BVTV'], target['BVTV'].values)[0])
        correlations['BVTV (p)'].append(pearsonr(results['BVTV'], target['BVTV'].values)[1])
        correlations['Tb.Th'].append(pearsonr(results['Trabecular thickness'], target['Tb.Th'].values)[0])
        correlations['Tb.Th (p)'].append(pearsonr(results['Trabecular thickness'], target['Tb.Th'].values)[1])
        correlations['Tb.Sp'].append(pearsonr(results['Trabecular separation'], target['Tb.Sp'].values)[0])
        correlations['Tb.Sp (p)'].append(pearsonr(results['Trabecular separation'], target['Tb.Sp'].values)[1])
        correlations['Tb.N'].append(pearsonr(results['Trabecular number'], target['Tb.N'].values)[0])
        correlations['Tb.N (p)'].append(pearsonr(results['Trabecular number'], target['Tb.N'].values)[1])

        # Save morphometric results to excel
        # Load existing file
        if os.path.isfile(savepath):
            book = load_workbook(savepath)
            writer = pd.ExcelWriter(savepath, engine='openpyxl', mode='a')
            writer.book = book
        else:
            writer = pd.ExcelWriter(savepath, engine='openpyxl')
        # Append new results
        df1 = pd.DataFrame(results)
        df1.to_excel(writer, sheet_name=snap_short)
        writer.save()

    # Save correlation results to excel
    # Load existing file
    book = load_workbook(savepath)
    writer = pd.ExcelWriter(savepath, engine='openpyxl', mode='a')
    writer.book = book
    df = pd.DataFrame(correlations)
    df.to_excel(writer, sheet_name='Correlations')
    writer.save()

    dur = time() - start
    completed = strftime(f'%Y_%m_%d_%H_%M')
    print(f'Analysis completed in {(dur % 3600) // 60} minutes, {dur % 60} seconds at time {completed}.')
